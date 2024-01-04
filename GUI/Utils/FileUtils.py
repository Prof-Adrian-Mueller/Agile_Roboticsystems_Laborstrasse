import datetime
import os

from PyQt6.QtGui import QPixmap
from PyQt6.QtWidgets import QFileDialog
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from GUI.Custom.CustomDialog import CustomDialog, ContentType
from openpyxl.drawing.image import Image

class FileUtils:
    @staticmethod
    def save_data_to_excel(parent, data_to_save, document_info):
        dialog = CustomDialog(parent)
        dialog.add_titlebar_name("Saving data to an Excel File")
        try:
            current_date = datetime.datetime.today().strftime('%Y-%m-%d')
            filename_suggestion = f"{current_date}_{document_info}.xlsx"

            filename, _ = QFileDialog.getSaveFileName(parent, "Save Excel File", filename_suggestion,
                                                      "Excel Files (*.xlsx)")
            if filename:
                if not filename.endswith('.xlsx'):
                    filename += '.xlsx'

                wb = Workbook()
                ws = wb.active
                image_saved = False

                for r_idx, row in enumerate(dataframe_to_rows(pd.DataFrame(data_to_save), index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        if isinstance(value, QPixmap):
                            # Save QPixmap as an image file
                            image_file = os.path.join(os.path.dirname(filename), f"image_{r_idx}_{c_idx}.png")
                            value.save(image_file)

                            # Insert the image into the Excel sheet
                            img = Image(image_file)
                            ws.add_image(img, f"{get_column_letter(c_idx)}{r_idx}")
                            image_saved = True
                        else:
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)

                            # Estimate and set row height
                            value_str = str(value) if value is not None else ''
                            column_width = ws.column_dimensions[cell.column_letter].width or 10  # Default to 10 if None
                            estimated_height = len(value_str) / column_width
                            current_height = ws.row_dimensions[cell.row].height or 15  # Default to 15 if None
                            ws.row_dimensions[cell.row].height = max(current_height, estimated_height)

                # Improved column width setting with padding
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = length + 2  # adding padding

                wb.save(filename)
                # Delete the saved images if any
                if image_saved:
                    for r_idx, row in enumerate(dataframe_to_rows(pd.DataFrame(data_to_save), index=False, header=True),
                                                1):
                        for c_idx, value in enumerate(row, 1):
                            if isinstance(value, QPixmap):
                                os.remove(os.path.join(os.path.dirname(filename), f"image_{r_idx}_{c_idx}.png"))

                dialog.addContent(f"Data saved to {filename}", ContentType.OUTPUT)
            else:
                dialog.addContent("Save operation cancelled.", ContentType.OUTPUT)

        except Exception as ex:
            dialog.addContent(f"Exception: {ex}", ContentType.OUTPUT)
        dialog.show()

